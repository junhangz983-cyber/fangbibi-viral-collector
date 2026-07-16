#!/usr/bin/env python3
"""Import Social Media Assistant video exports into a Feishu Base."""

from __future__ import annotations

import argparse
import json
import os
import subprocess
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_COLUMNS = {
    "视频ID",
    "视频链接",
    "点赞量",
    "评论量",
    "分享量",
    "发布时间",
    "视频时长",
    "达人昵称",
}

FEISHU_FIELDS = [
    "视频标题",
    "链接",
    "平台",
    "达人昵称",
    "发布时间",
    "点赞",
    "评论",
    "分享",
    "视频时长",
    "视频 ID",
    "爆点判断",
    "入库日期",
]


def run_lark_cli(command: list[str]) -> subprocess.CompletedProcess[str]:
    environment = os.environ | {
        "LARKSUITE_CLI_NO_UPDATE_NOTIFIER": "1",
        "LARKSUITE_CLI_NO_SKILLS_NOTIFIER": "1",
    }
    return subprocess.run(command, check=False, capture_output=True, text=True, env=environment)


def load_export(path: Path, min_likes: int, imported_at: str) -> list[list[Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    missing_columns = REQUIRED_COLUMNS - set(headers)
    if missing_columns:
        missing = "、".join(sorted(missing_columns))
        raise ValueError(f"{path.name} 不是支持的视频导出，缺少：{missing}")

    rows_by_link: dict[str, list[Any]] = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, values))
        link = str(record.get("视频链接") or "").strip()
        likes = int(record.get("点赞量") or 0)
        if not link or likes < min_likes:
            continue

        video_id = str(record["视频ID"])
        creator = str(record.get("达人昵称") or "未知达人")
        published_at = record.get("发布时间")
        if isinstance(published_at, datetime):
            published_at = published_at.strftime("%Y-%m-%d %H:%M:%S")
        comments = int(record.get("评论量") or 0)
        shares = int(record.get("分享量") or 0)
        rows_by_link[link] = [
            f"待补标题｜{creator}｜{video_id}",
            link,
            "抖音",
            creator,
            published_at,
            likes,
            comments,
            shares,
            record.get("视频时长"),
            video_id,
            f"数据爆款：点赞 {likes:,}，评论 {comments:,}，分享 {shares:,}。",
            imported_at,
        ]
    return list(rows_by_link.values())


def existing_links(base_token: str, table_id: str) -> set[str]:
    links: set[str] = set()
    offset = 0
    while True:
        command = [
            "lark-cli",
            "base",
            "+record-list",
            "--as",
            "user",
            "--base-token",
            base_token,
            "--table-id",
            table_id,
            "--field-id",
            "链接",
            "--offset",
            str(offset),
            "--limit",
            "200",
            "--format",
            "json",
        ]
        result = run_lark_cli(command)
        if result.returncode:
            raise RuntimeError(result.stderr.strip() or result.stdout.strip())
        payload = json.loads(result.stdout).get("data", {})
        for record in payload.get("data", []):
            if record and isinstance(record[0], str):
                links.add(record[0])
        if not payload.get("has_more"):
            return links
        offset += 200


def write_batches(rows: list[list[Any]], base_token: str, table_id: str) -> int:
    for start in range(0, len(rows), 200):
        payload = json.dumps(
            {"fields": FEISHU_FIELDS, "rows": rows[start : start + 200]},
            ensure_ascii=False,
        )
        command = [
            "lark-cli",
            "base",
            "+record-batch-create",
            "--as",
            "user",
            "--base-token",
            base_token,
            "--table-id",
            table_id,
            "--json",
            payload,
        ]
        result = run_lark_cli(command)
        if result.returncode:
            raise RuntimeError(result.stderr.strip() or result.stdout.strip())
    return len(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="导入社媒助手抖音视频 Excel 到飞书多维表格")
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--base-token", required=True)
    parser.add_argument("--table-id", required=True)
    parser.add_argument("--min-likes", type=int, default=500)
    parser.add_argument("--imported-at", default=datetime.now().strftime("%Y-%m-%d"))
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    rows = load_export(args.input, args.min_likes, args.imported_at)
    known_links = existing_links(args.base_token, args.table_id)
    rows = [row for row in rows if row[1] not in known_links]
    if args.dry_run:
        print(json.dumps({"qualified": len(rows), "written": 0}, ensure_ascii=False))
        return 0

    written = write_batches(rows, args.base_token, args.table_id)
    print(json.dumps({"qualified": len(rows), "written": written}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
